import pandas as pd
import os
import json
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

# Set up command line argument parsing for the config file
parser = argparse.ArgumentParser(description='Process element data with JSON configuration file.')
parser.add_argument('--config', type=str, default='config.json',
                    help='Path to JSON configuration file (default: config.json)')
                    
# Parse command line arguments
args = parser.parse_args()
config_file = args.config

# Default configuration
config = {
    "low_cut_off_metal": 38,
    "high_cut_off_silicate": 8,
    "files": ["fe_in_fe_Count.xlsx", "mg_in_fe_Count.xlsx", "si_in_fe_Count.xlsx", 
              "o_in_fe_Count.xlsx", "n_in_fe_Count.xlsx", "h_in_fe_Count.xlsx"],
    "atomic_mass": {
        "fe": 55.845,
        "mg": 24.305,
        "si": 28.0855,
        "o": 15.9994,
        "n": 14.0067,
        "h": 1.00784
    },
    "column_widths": {
        "composition": {
            "A": 15, "B": 20, "C": 15, "D": 15, "E": 20,
            "F": 20, "G": 20, "H": 20, "I": 20, "J": 20, "K": 20
        },
        "final": {
            "A": 15, "B": 15, "C": 15, "D": 15, "E": 15,
            "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15
        }
    }
}

# Try to load configuration from JSON file
try:
    if os.path.exists(config_file):
        with open(config_file, 'r') as f:
            loaded_config = json.load(f)
            # Update default config with loaded values
            config.update(loaded_config)
        print(f"Successfully loaded configuration from {config_file}")
    else:
        print(f"Config file {config_file} not found. Using default configuration.")
        # Create a sample config file for future reference
        with open('sample_config.json', 'w') as f:
            json.dump(config, f, indent=4)
        print(f"Created sample configuration file: sample_config.json")
except Exception as e:
    print(f"Error loading configuration: {e}")
    print("Using default configuration.")

# Extract configuration values
low_cut_off_metal = config["low_cut_off_metal"]
high_cut_off_silicate = config["high_cut_off_silicate"]
files = config["files"]
atomic_mass_mapper = config["atomic_mass"]
column_widths_comp = config["column_widths"]["composition"]
column_widths_final = config["column_widths"]["final"]

print(f"Using low_cut_off_metal = {low_cut_off_metal}")
print(f"Using high_cut_off_silicate = {high_cut_off_silicate}")
print(f"Processing files: {files}")

# Read atom counts
atom_count_list = []
atom_names = []
fe_atom_count = 1  # Default value

# Try to read the Fe atom count
try:
    with open("fe_count.txt", "r") as f:
        fe_atom_count = int(f.read().strip())
except Exception as e:
    print(f"Error reading fe_count.txt: {e}")

# Read atom counts for each element
for file in files:
    if os.path.exists(file):
        atom_name = file.split("_")[0]
        atom_names.append(atom_name)
        
        # Read atom count from corresponding text file
        try:
            with open(f"{atom_name}_count.txt", "r") as f:
                atom_count_list.append(int(f.read().strip()))
        except Exception as e:
            print(f"Error reading {atom_name}_count.txt: {e}")
            atom_count_list.append(0)  # Default value if file can't be read

# Define function to process a row (whether Average or Block)
def process_row(row_name):
    list_data = []
    index_data = ["Bin Count"]
    available_atoms = []
    
    # First pass: determine which files have the requested row and collect atom names
    for file in files:
        if os.path.exists(file):
            atom_name = file.split("_")[0]
            atom_df = pd.read_excel(file, index_col=0)
            
            if row_name in atom_df.index:
                available_atoms.append(atom_name)
    
    print(f"Available atoms for {row_name}: {available_atoms}")
    
    # Second pass: process each file
    first_file_processed = False
    for file in files:
        if not os.path.exists(file):
            continue
            
        atom_name = file.split("_")[0]
        
        # Load the Excel file and extract the specified row's data
        atom_df = pd.read_excel(file, index_col=0)
        
        if row_name not in atom_df.index:
            print(f"Warning: {row_name} not found in {file}")
            continue
                
        atom_data = atom_df.loc[row_name].values
        
        # For the first available file, initialize Bin Count
        if not first_file_processed:
            bin_count = list(range(len(atom_data)))
            list_data.append(bin_count)   # Bin Count
            first_file_processed = True
        
        # Round atom_data to 2 decimal places
        atom_data = np.round(atom_data, 2)
        list_data.append(atom_data)  # Atom data for each element
        index_data.append(f"{atom_name} atoms")  # Update column names
    
    # Create the final DataFrame and transpose it
    final_df = pd.DataFrame(list_data, index=index_data).T  # Transpose rows to columns
    
    # Calculate percentages for internal use only (not to be saved)
    percentages_for_calculation = {}
    for i, atom_name in enumerate(atom_names):
        if f"{atom_name} atoms" in final_df.columns and i < len(atom_count_list):
            if atom_count_list[i] > 0:  # Prevent division by zero
                percentages_for_calculation[atom_name] = round(100 * final_df[f"{atom_name} atoms"]  / atom_count_list[i], 2)
            else:
                percentages_for_calculation[atom_name] = pd.Series(0, index=final_df.index)
    
    # Filter for Metal region (using command line argument for low cutoff)
    filtered_metal_df = final_df[(final_df["Bin Count"] >= low_cut_off_metal)]
    
    # Filter for Silicate region (using command line argument for high cutoff)
    filtered_silicate_df = final_df[final_df["Bin Count"] <= high_cut_off_silicate]
    
    # Construct composition dictionary with placeholders
    composition = {
        "atom_name": [],
        "total_composition": [],
        "atomic_mass": [],
        "metal_percent": [],
        "silicate_percent": []
    }
    
    # Print debug info
    print(f"Processing composition for {row_name}")
    
    # Add data for all elements
    for i, atom_name in enumerate(atom_names):
        if atom_name not in percentages_for_calculation:
            print(f"Skipping {atom_name} for composition as it's not in percentage calculations")
            continue  # Skip if data not available
            
        if i >= len(atom_count_list):
            print(f"Skipping {atom_name} for composition as index {i} is out of bounds for atom_count_list")
            continue  # Skip if index out of bounds
        
        composition["atom_name"].append(atom_name)
        composition["atomic_mass"].append(round(atomic_mass_mapper.get(atom_name, 0), 2))
        composition["total_composition"].append(atom_count_list[i])
        
        # Calculate metal and silicate percentages
        metal_percent = percentages_for_calculation[atom_name].loc[filtered_metal_df.index].sum() if not filtered_metal_df.empty else 0
        silicate_percent = percentages_for_calculation[atom_name].loc[filtered_silicate_df.index].sum() if not filtered_silicate_df.empty else 0
        
        # Round percentages to 2 decimal places
        metal_percent = round(metal_percent, 2)
        silicate_percent = round(silicate_percent, 2)
        
        print(f"  {atom_name} - Metal: {metal_percent}, Silicate: {silicate_percent}")
        
        composition["metal_percent"].append(metal_percent)
        composition["silicate_percent"].append(silicate_percent)
    
    # Create composition DataFrame
    composition_df = pd.DataFrame(composition)
    
    # Calculate atom counts and masses
    composition_df["metal_atom_count"] = round(composition_df["total_composition"] * composition_df["metal_percent"] / 100, 2)
    composition_df["silicate_atom_count"] = round(composition_df["total_composition"] * composition_df["silicate_percent"] / 100, 2)
    
    composition_df["metal_mass"] = round(composition_df["atomic_mass"] * composition_df["metal_atom_count"], 2)
    composition_df["silicate_mass"] = round(composition_df["atomic_mass"] * composition_df["silicate_atom_count"], 2)
    
    # Calculate total masses and weight percentages
    metal_mass_total = composition_df["metal_mass"].sum()
    silicate_mass_total = composition_df["silicate_mass"].sum()
    
    # Prevent division by zero
    if metal_mass_total > 0:
        composition_df["metal_weight_percent"] = round((composition_df["metal_mass"] / metal_mass_total)*100, 2)
    else:
        composition_df["metal_weight_percent"] = 0
        
    if silicate_mass_total > 0:
        composition_df["silicate_weight_percent"] = round((composition_df["silicate_mass"] / silicate_mass_total)*100, 2)
    else:
        composition_df["silicate_weight_percent"] = 0
    
    # Add row_name as a column to identify the source of this data
    composition_df.insert(0, "block_name", row_name)
    
    return final_df, composition_df

# Function to handle NaN values when converting to JSON
def convert_nan_to_null(obj):
    if isinstance(obj, float) and pd.isna(obj):
        return None
    return obj

# Get list of all row names to process (Blocks + Average)
row_names = []

# Read first file to get row names
if files and os.path.exists(files[0]):
    df = pd.read_excel(files[0], index_col=0)
    for idx in df.index:
        if idx == "Average" or str(idx).startswith("Block"):
            row_names.append(idx)

print(f"Found {len(row_names)} rows to process: {row_names}")

# Try to load cell values from temp.xlsx
cell_values = {}
try:
    if os.path.exists('temp.xlsx'):
        temp_df = pd.read_excel('temp.xlsx')
        # Check for all possible cell columns
        for element in ['Fe', 'Mg', 'Si', 'O', 'H', 'N']:
            cell_col = f"{element} cell"
            if cell_col in temp_df.columns:
                # Round cell values to 2 decimal places
                cell_values[cell_col] = [round(val, 2) if isinstance(val, float) else val 
                                        for val in temp_df[cell_col].tolist()]
                print(f"Loaded {cell_col} values from temp.xlsx, length: {len(cell_values[cell_col])}")
except Exception as e:
    print(f"Error loading cell values from temp.xlsx: {e}")

# Create output filename that includes the cutoff values
output_filename = f"results.xlsx"

# Create output directory for JSON files if it doesn't exist
json_dir = "json_output"
os.makedirs(json_dir, exist_ok=True)

# Create containers for all final and composition data
all_final_dfs = {}
all_composition_dfs = []

# Process each row name
for row_name in row_names:
    print(f"Processing {row_name}...")
    final_df, composition_df = process_row(row_name)
    
    # Remove percentage columns before saving - ALWAYS DO THIS FOR ALL BLOCKS
    columns_to_keep = [col for col in final_df.columns if "(%)" not in col]
    final_df_no_percentages = final_df[columns_to_keep]
    
    # For Average, add cell columns from temp.xlsx if available
    if row_name == "Average" and cell_values:
        for cell_col, values in cell_values.items():
            if len(values) == len(final_df_no_percentages):
                final_df_no_percentages[cell_col] = values
            else:
                print(f"Warning: {cell_col} length mismatch. Expected {len(final_df_no_percentages)}, got {len(values)}")
    
    # Round all float values in the final DataFrame to 2 decimal places
    for col in final_df_no_percentages.columns:
        if final_df_no_percentages[col].dtype == float:
            final_df_no_percentages[col] = final_df_no_percentages[col].round(2)
    
    # Store the final dataframe
    all_final_dfs[row_name] = final_df_no_percentages
    
    # Round all float values in the composition DataFrame to 2 decimal places
    for col in composition_df.columns:
        if composition_df[col].dtype == float:
            composition_df[col] = composition_df[col].round(2)
    
    # Store the composition dataframe to be combined later
    all_composition_dfs.append(composition_df)
    
    # Export Final dataframe to JSON with pretty printing
    json_filename = os.path.join(json_dir, f"Final_{row_name.replace(' ', '_')}.json")
    
    # Convert dataframe to dictionary and handle NaN values
    final_data = final_df_no_percentages.to_dict(orient='records')
    
    # Write with indentation for pretty printing
    with open(json_filename, 'w') as f:
        json.dump(final_data, f, indent=4, default=convert_nan_to_null)
        
    print(f"Exported Final_{row_name} data to {json_filename}")

# Create a new Excel writer
with pd.ExcelWriter(output_filename, engine="xlsxwriter") as writer:
    # Save each final DataFrame to its own sheet
    for row_name, df in all_final_dfs.items():
        # Convert row_name to a valid sheet name
        sheet_name_base = row_name.replace(" ", "_")
        final_sheet_name = f"Final_{sheet_name_base}"
        if len(final_sheet_name) > 31:  # Excel sheet name length limit
            final_sheet_name = final_sheet_name[:31]
        df.to_excel(writer, sheet_name=final_sheet_name, index=False)
    
    # Create concatenated composition dataframe with empty rows between blocks
    all_comp_data = []
    for i, comp_df in enumerate(all_composition_dfs):
        all_comp_data.append(comp_df)
        if i < len(all_composition_dfs) - 1:
            # Add empty row unless this is the last block
            # Create an empty dataframe with the same columns
            empty_df = pd.DataFrame(columns=comp_df.columns)
            all_comp_data.append(empty_df)
    
    # Concatenate all the pieces
    combined_comp_df = pd.concat(all_comp_data, ignore_index=True)
    
    # Save combined composition DataFrame to a single sheet
    combined_comp_df.to_excel(writer, sheet_name="Composition", index=False)

# Export all composition data to a single JSON file
comp_json_filename = os.path.join(json_dir, "Composition.json")
composition_by_block = {}

# Group composition data by block for easier access in the JSON
for comp_df in all_composition_dfs:
    block_name = comp_df['block_name'].iloc[0]
    # Convert to records format and remove the block_name column from each record
    # since it's redundant as a key in the dictionary
    block_data = comp_df.drop('block_name', axis=1).to_dict(orient='records')
    composition_by_block[block_name] = block_data

# Write the composition data to JSON with indentation for pretty printing
with open(comp_json_filename, 'w') as f:
    json.dump(composition_by_block, f, indent=4, default=convert_nan_to_null)
print(f"Exported composition data to {comp_json_filename}")

print("Processing complete!")

# Load the workbook to adjust column widths and apply formatting
wb = load_workbook(output_filename)

# Define colors for each block
block_colors = {
    "Average": "E0FFFF",  # Light Cyan
    "Block 1": "FFD700",  # Gold
    "Block 2": "98FB98",  # Pale Green
    "Block 3": "FFA07A",  # Light Salmon
    "Block 4": "B0E0E6",  # Powder Blue
    "Block 5": "FFDAB9",  # Peach Puff
    "Block 6": "D8BFD8",  # Thistle
    "Block 7": "87CEFA",  # Light Sky Blue
    "Block 8": "FFFACD",  # Lemon Chiffon
    "Block 9": "F0FFF0",  # Honeydew
    "Block 10": "F5F5DC"  # Beige
}

# Default color for any blocks not in the dictionary
default_color = "F0F0F0"  # Light Gray

# Apply formatting to the Composition sheet
if "Composition" in wb.sheetnames:
    ws = wb["Composition"]
    
    # Create a thin border style
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Start from row 2 (after header)
    current_row = 2
    current_block = None
    block_start = None
    
    # Scan through all rows and color them based on block
    while current_row <= ws.max_row:
        cell_value = ws.cell(row=current_row, column=1).value
        
        # Check if this is an empty row (separator)
        if cell_value is None:
            # If we were processing a block, finalize it with borders
            if current_block and block_start:
                block_end = current_row - 1
                color = block_colors.get(current_block, default_color)
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Apply fill and borders to the block
                for row in range(block_start, block_end + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill
                        cell.border = thin_border
                        
                # Reset for next block
                current_block = None
                block_start = None
            
            current_row += 1
            continue
        
        # Check if this is a new block
        if isinstance(cell_value, str) and (cell_value == "Average" or cell_value.startswith("Block")):
            if current_block != cell_value:
                # If we were processing a previous block, finalize it
                if current_block and block_start:
                    block_end = current_row - 1
                    color = block_colors.get(current_block, default_color)
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # Apply fill and borders to the block
                    for row in range(block_start, block_end + 1):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = fill
                            cell.border = thin_border
                
                # Start new block
                current_block = cell_value
                block_start = current_row
        
        current_row += 1
    
    # Handle the last block
    if current_block and block_start:
        block_end = current_row - 1
        color = block_colors.get(current_block, default_color)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Apply fill and borders to the block
        for row in range(block_start, block_end + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border

# Apply column widths to all sheets
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Select appropriate column widths based on sheet name
    if sheet_name == "Composition":
        column_widths = column_widths_comp
    else:
        column_widths = column_widths_final
        
    # Apply column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

# Save the modified workbook
wb.save(output_filename)

print(f"All processing complete! Results saved to '{output_filename}'")
print(f"JSON data saved to '{json_dir}' directory")

# Cell Frequency Percent code

from openpyxl import load_workbook

# Load temp.xlsx and drop unnamed columns
temp_df = pd.read_excel("temp.xlsx", sheet_name=None)  # Load all sheets

# Drop unnamed columns if they exist
for sheet in temp_df:
    temp_df[sheet] = temp_df[sheet].dropna(axis=1, how='all')  # Drop fully empty columns

# Load results.xlsx and get existing sheets
with pd.ExcelFile("results.xlsx") as xls:
    sheet_names = xls.sheet_names

new_sheet_name = "Cell Frequency Percent"

# Write to results.xlsx while preserving existing sheets
with pd.ExcelWriter("results.xlsx", mode="a", engine="openpyxl") as writer:
    for sheet in temp_df:
        temp_df[sheet].to_excel(writer, sheet_name=new_sheet_name, index=False)  # Ensure index=False

# Load the workbook and access the new sheet
wb = load_workbook("results.xlsx")
ws = wb[new_sheet_name]

# Set column width to 15
for col in ws.columns:
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 15

wb.save("results.xlsx")

# Save json output

if not os.path.exists("json_output"):
    print("Creating json_output folder")
    os.makedirs("json_output")

# Save temp_df to JSON in the json_output folder
json_data = {}
for sheet in temp_df:
    # Convert DataFrame to dict records
    json_data[sheet] = temp_df[sheet].to_dict(orient="records")

# Save the JSON file
json_filepath = os.path.join("json_output", "cellFrequencyPercent.json")
with open(json_filepath, "w") as json_file:
    json.dump(json_data, json_file, indent=2)

print(f"Cell Frequency Percent data saved to {json_filepath}")
